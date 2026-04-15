"""
Importación masiva de precios / tipo de precio / unidad desde plantilla Excel
(misma estructura que TransactionLinesExcelPanel: fila 1 códigos, fila 2 descripciones, datos desde fila 3).
"""
from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.authentication import TokenAuthentication
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from appinventory.models import Product, ProductPrice, PriceType, UnitOfMeasure

logger = logging.getLogger(__name__)


def _norm_header(cell) -> str:
    return (
        str(cell if cell is not None else "")
        .strip()
        .lower()
        .replace(" ", "_")
    )


def _build_col_map(header_row: tuple) -> dict:
    headers = [_norm_header(c) for c in header_row]

    def find_col(*aliases: str) -> int:
        for a in aliases:
            try:
                return headers.index(a)
            except ValueError:
                continue
        return -1

    return {
        "pid": find_col("product_id", "productid", "id_product"),
        "unit": find_col("unit_code", "unit", "uom"),
        "price": find_col("unit_price", "price"),
        "pt": find_col("price_type_name", "price_type", "pricetype"),
    }


def _to_int_pid(raw) -> int | None:
    if raw is None or str(raw).strip() == "":
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    s = str(raw).strip()
    try:
        return int(Decimal(s))
    except (InvalidOperation, ValueError, TypeError):
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None


def _to_decimal_price(raw) -> Decimal | None:
    if raw is None or str(raw).strip() == "":
        return None
    try:
        return Decimal(str(raw).strip())
    except InvalidOperation:
        return None


class ProductPricesBulkImportAPIView(APIView):
    """
    POST multipart: field `file` = .xlsx (misma plantilla que líneas de transacción).

    Por cada fila: actualiza `Product.unit_default` según unit_code (si viene),
    e inserta o actualiza un `ProductPrice` de venta (is_sale=True, is_purchase=False)
    sin ventana de vigencia (valid_from/valid_until null) para el par (price_type, unit).
    """

    authentication_classes = [TokenAuthentication]
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if not request.user.has_perm("appinventory.change_product"):
            return Response(
                {"detail": "No permission to change products."},
                status=status.HTTP_403_FORBIDDEN,
            )
        if not request.user.has_perm("appinventory.add_productprice") or not request.user.has_perm(
            "appinventory.change_productprice"
        ):
            return Response(
                {"detail": "No permission to add/change product prices."},
                status=status.HTTP_403_FORBIDDEN,
            )

        upload = request.FILES.get("file")
        if not upload:
            return Response(
                {"detail": 'Missing file field "file".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        name = (getattr(upload, "name", "") or "").lower()
        if not name.endswith(".xlsx"):
            return Response(
                {"detail": "Only .xlsx files are supported (use the downloaded template)."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            data = upload.read()
            wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        except Exception as e:
            logger.exception("bulk prices: could not read workbook")
            return Response(
                {"detail": f"Could not read Excel file: {e}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            return Response(
                {
                    "detail": "File must have a header row, a description row, and at least one data row."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        col = _build_col_map(rows[0])
        if col["pid"] < 0:
            return Response(
                {"detail": 'Missing required column "product_id" in first row.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Caches
        units_all = list(UnitOfMeasure.objects.filter(is_active=True).only("id", "code", "name"))
        unit_by_code = {}
        unit_by_name = {}
        for u in units_all:
            if u.code:
                unit_by_code[str(u.code).strip().lower()] = u
            if u.name:
                unit_by_name[str(u.name).strip().lower()] = u

        pt_all = list(PriceType.objects.filter(is_active=True).only("id", "name"))
        pt_by_name = {str(p.name).strip().lower(): p for p in pt_all}

        errors: list[dict] = []
        # Filas válidas: excel_row (1-based), product_id, unit_id, price_type_id, price, set_unit_default
        planned: list[dict] = []

        for i, row in enumerate(rows[2:], start=3):
            pid = _to_int_pid(row[col["pid"]] if col["pid"] >= 0 else None)
            if pid is None:
                continue

            raw_pt = row[col["pt"]] if col["pt"] >= 0 else None
            pt_name = str(raw_pt).strip() if raw_pt is not None else ""
            if not pt_name:
                errors.append({"row": i, "message": "price_type_name is required."})
                continue

            pt_obj = pt_by_name.get(pt_name.lower())
            if not pt_obj:
                errors.append(
                    {
                        "row": i,
                        "message": f'Unknown price type "{raw_pt}". Use exact name from Price Types.',
                    }
                )
                continue

            raw_unit = row[col["unit"]] if col["unit"] >= 0 else None
            unit_str = str(raw_unit).strip() if raw_unit is not None else ""

            raw_price = row[col["price"]] if col["price"] >= 0 else None
            price_dec = _to_decimal_price(raw_price)
            if price_dec is None:
                errors.append({"row": i, "message": "unit_price is missing or invalid."})
                continue
            if price_dec < 0:
                errors.append({"row": i, "message": "unit_price cannot be negative."})
                continue

            planned.append(
                {
                    "excel_row": i,
                    "product_id": pid,
                    "price_type_id": pt_obj.id,
                    "unit_str": unit_str,
                    "price": price_dec,
                }
            )

        if not planned and not errors:
            return Response(
                {"detail": "No data rows with product_id found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        product_ids = {p["product_id"] for p in planned}
        products = {p.id: p for p in Product.objects.filter(id__in=product_ids)}
        missing_products = product_ids - set(products.keys())
        if missing_products:
            for p in planned:
                if p["product_id"] in missing_products:
                    errors.append(
                        {
                            "row": p["excel_row"],
                            "message": f"Product id {p['product_id']} does not exist.",
                        }
                    )
            planned = [p for p in planned if p["product_id"] not in missing_products]

        # Resolver unidad por fila
        resolved: list[dict] = []
        for p in planned:
            prod = products[p["product_id"]]
            unit_str = p["unit_str"]
            unit_obj = None
            if unit_str:
                unit_obj = unit_by_code.get(unit_str.lower()) or unit_by_name.get(
                    unit_str.lower()
                )
                if not unit_obj:
                    errors.append(
                        {
                            "row": p["excel_row"],
                            "message": f'Unknown unit (code or name): "{unit_str}".',
                        }
                    )
                    continue
            else:
                unit_obj = prod.unit_default
                if not unit_obj:
                    errors.append(
                        {
                            "row": p["excel_row"],
                            "message": "unit_code is empty and product has no default unit.",
                        }
                    )
                    continue

            resolved.append(
                {
                    "excel_row": p["excel_row"],
                    "product_id": p["product_id"],
                    "price_type_id": p["price_type_id"],
                    "unit_id": unit_obj.id,
                    "price": p["price"],
                    "set_unit_default": bool(unit_str),
                }
            )

        if not resolved:
            return Response(
                {
                    "detail": "No rows could be applied.",
                    "errors": errors,
                    "created": 0,
                    "updated": 0,
                    "unit_default_updated": 0,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Misma combinación (producto, tipo precio, unidad): gana la última fila del Excel (por número de fila)
        _dedup: dict[tuple, dict] = {}
        for r in sorted(resolved, key=lambda x: x["excel_row"]):
            key = (r["product_id"], r["price_type_id"], r["unit_id"])
            _dedup[key] = r
        resolved = list(_dedup.values())

        # Upsert ProductPrice (venta, sin vigencia)
        created = 0
        updated = 0
        unit_default_updated = 0

        with transaction.atomic():
            # Actualizar unit_default en productos indicados
            products_to_save = []
            seen_pid_unit = {}
            for r in sorted(resolved, key=lambda x: x["excel_row"]):
                if r["set_unit_default"]:
                    seen_pid_unit[r["product_id"]] = r["unit_id"]
            for pid, uid in seen_pid_unit.items():
                pr = products.get(pid)
                if pr and pr.unit_default_id != uid:
                    pr.unit_default_id = uid
                    products_to_save.append(pr)
            if products_to_save:
                Product.objects.bulk_update(products_to_save, ["unit_default_id"])
                unit_default_updated = len(products_to_save)

            product_ids_q = {r["product_id"] for r in resolved}
            existing_qs = ProductPrice.objects.filter(
                product_id__in=product_ids_q,
                is_purchase=False,
                is_sale=True,
                valid_from__isnull=True,
                valid_until__isnull=True,
            )
            existing_map = {}
            for ep in existing_qs:
                key = (ep.product_id, ep.price_type_id, ep.unit_id)
                existing_map[key] = ep

            to_create = []
            to_update = []

            for r in resolved:
                key = (r["product_id"], r["price_type_id"], r["unit_id"])
                price_val = r["price"].quantize(Decimal("0.01"))
                if key in existing_map:
                    obj = existing_map[key]
                    if obj.price != price_val:
                        obj.price = price_val
                        obj.is_active = True
                        to_update.append(obj)
                else:
                    to_create.append(
                        ProductPrice(
                            product_id=r["product_id"],
                            price_type_id=r["price_type_id"],
                            unit_id=r["unit_id"],
                            price=price_val,
                            is_purchase=False,
                            is_sale=True,
                            is_default=False,
                            is_active=True,
                            valid_from=None,
                            valid_until=None,
                        )
                    )

            if to_update:
                ProductPrice.objects.bulk_update(to_update, ["price", "is_active"])
                updated = len(to_update)

            if to_create:
                ProductPrice.objects.bulk_create(to_create)
                created = len(to_create)

            # Si el producto no tiene precio de venta marcado como default, asignar uno
            affected_pids = {r["product_id"] for r in resolved}
            for pid in affected_pids:
                sale_prices = list(
                    ProductPrice.objects.filter(
                        product_id=pid,
                        is_sale=True,
                        is_active=True,
                    )
                )
                if not sale_prices:
                    continue
                if sum(1 for x in sale_prices if x.is_default) == 0:
                    chosen = min(sale_prices, key=lambda x: x.id)
                    ProductPrice.objects.filter(
                        product_id=pid, is_default=True, is_sale=True
                    ).update(is_default=False)
                    ProductPrice.objects.filter(pk=chosen.pk).update(is_default=True)

        return Response(
            {
                "created": created,
                "updated": updated,
                "unit_default_updated": unit_default_updated,
                "rows_applied": len(resolved),
                "errors": errors,
            },
            status=status.HTTP_200_OK,
        )
