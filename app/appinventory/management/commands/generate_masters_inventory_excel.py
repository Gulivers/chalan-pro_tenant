from django.core.management.base import BaseCommand
from django.conf import settings
import json
import os
from pathlib import Path


class Command(BaseCommand):
    help = 'Generate Excel file from masters_inventory.json fixture for download'

    def add_arguments(self, parser):
        parser.add_argument(
            '--output-dir',
            type=str,
            default=None,
            help='Output directory for Excel file (default: appinventory/static/appinventory/)'
        )

    def handle(self, *args, **options):
        # Ruta del fixture JSON
        fixture_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
            'appinventory',
            'fixtures',
            'masters_inventory.json'
        )
        
        if not os.path.exists(fixture_path):
            self.stdout.write(self.style.ERROR(f'[ERROR] Fixture file not found: {fixture_path}'))
            return
        
        # Directorio de salida
        if options['output_dir']:
            output_dir = options['output_dir']
        else:
            # Usar static/appinventory/
            output_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))),
                'appinventory',
                'static',
                'appinventory'
            )
        
        # Crear directorio si no existe
        os.makedirs(output_dir, exist_ok=True)
        
        output_file = os.path.join(output_dir, 'masters_inventory.xlsx')
        
        try:
            # Cargar datos del JSON
            self.stdout.write(self.style.SUCCESS(f'[INFO] Loading data from {fixture_path}...'))
            with open(fixture_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Mapeo de nombres de modelo en minúsculas a PascalCase
            model_name_mapping = {
                'unitcategory': 'UnitCategory',
                'unitofmeasure': 'UnitOfMeasure',
                'warehouse': 'Warehouse',
                'productcategory': 'ProductCategory',
                'productbrand': 'ProductBrand',
                'pricetype': 'PriceType',
                'product': 'Product',
                'productprice': 'ProductPrice'
            }
            
            # Organizar datos por modelo
            model_data = {
                'UnitCategory': [],
                'UnitOfMeasure': [],
                'Warehouse': [],
                'ProductCategory': [],
                'ProductBrand': [],
                'PriceType': [],
                'Product': [],
                'ProductPrice': []
            }
            
            for item in data:
                # Extraer nombre del modelo (ej: "appinventory.unitcategory" -> "unitcategory")
                raw_model_name = item.get('model', '').split('.')[-1].lower() if '.' in item.get('model', '') else item.get('model', '').lower()
                # Convertir a PascalCase usando el mapeo
                model_name = model_name_mapping.get(raw_model_name)
                
                if model_name and model_name in model_data:
                    model_data[model_name].append({
                        'pk': item.get('pk'),
                        **item.get('fields', {})
                    })
            
            # Generar Excel
            self.stdout.write(self.style.SUCCESS('[INFO] Generating Excel file...'))
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            
            # Mapeo de nombres de modelo a nombres de hoja
            sheet_names = {
                'UnitCategory': 'Categorías de Unidades',
                'UnitOfMeasure': 'Unidades de Medida',
                'Warehouse': 'Almacenes',
                'ProductCategory': 'Categorías de Productos',
                'ProductBrand': 'Marcas',
                'PriceType': 'Tipos de Precio',
                'Product': 'Productos',
                'ProductPrice': 'Precios de Productos'
            }
            
            sheets_created = 0
            for model_name, items in model_data.items():
                if not items:
                    continue
                
                # Remover la hoja por defecto solo después de crear la primera hoja
                if sheets_created == 0 and wb.active:
                    wb.remove(wb.active)
                
                ws = wb.create_sheet(title=sheet_names.get(model_name, model_name))
                sheets_created += 1
                
                # Obtener todas las claves únicas de todos los items
                all_keys = set()
                for item in items:
                    all_keys.update(key for key in item.keys() if key != 'pk')
                
                headers = ['ID'] + sorted(all_keys)
                
                # Estilo para encabezados
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                
                # Escribir encabezados
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Escribir datos
                for row_idx, item in enumerate(items, 2):
                    ws.cell(row=row_idx, column=1, value=item.get('pk', ''))
                    for col_idx, header in enumerate(headers[1:], 2):
                        value = item.get(header, '')
                        # Formatear valores booleanos
                        if isinstance(value, bool):
                            value = 'Sí' if value else 'No'
                        # Formatear objetos complejos como texto
                        elif isinstance(value, (dict, list)):
                            value = str(value)
                        ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Auto-ajustar anchos de columna
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            wb.save(output_file)
            self.stdout.write(self.style.SUCCESS(f'[SUCCESS] Excel file generated: {output_file}'))
            
        except ImportError:
            self.stdout.write(self.style.ERROR('[ERROR] openpyxl is not installed. Install it with: pip install openpyxl'))
        except Exception as e:
            self.stdout.write(self.style.ERROR(f'[ERROR] Error generating Excel: {str(e)}'))
            raise
